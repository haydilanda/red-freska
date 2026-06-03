import csv
import io
from fastapi import APIRouter, HTTPException, Depends, UploadFile, File
from supabase import create_client
from models import Tendencia, TendenciaCreate
from routers.auth import get_current_user
from typing import List, Optional
import os
import openpyxl

router = APIRouter(prefix="/tendencias", tags=["tendencias"])


def get_supabase():
    return create_client(
        os.getenv("SUPABASE_URL"),
        os.getenv("SUPABASE_SERVICE_KEY"),
    )


def require_admin(user=Depends(get_current_user)):
    if user["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo admins")
    return user


@router.get("/", response_model=List[Tendencia])
async def listar_tendencias(
    estado: Optional[str] = None,
    user=Depends(get_current_user),
):
    sb = get_supabase()
    query = sb.table("tendencias").select("*").order("creada_at", desc=True)

    # Clientes solo ven tendencias publicadas
    if user["rol"] == "cliente":
        query = query.eq("estado", "publicada")
    elif estado:
        query = query.eq("estado", estado)

    res = query.execute()
    return res.data


@router.get("/{tendencia_id}", response_model=Tendencia)
async def obtener_tendencia(tendencia_id: str, user=Depends(get_current_user)):
    sb = get_supabase()
    res = (
        sb.table("tendencias")
        .select("*")
        .eq("id", tendencia_id)
        .single()
        .execute()
    )
    if not res.data:
        raise HTTPException(status_code=404, detail="Tendencia no encontrada")

    # Cliente no puede ver tendencias en revisión
    if user["rol"] == "cliente" and res.data["estado"] != "publicada":
        raise HTTPException(status_code=403, detail="Acceso denegado")

    return res.data


@router.post("/", response_model=Tendencia)
async def crear_tendencia(t: TendenciaCreate, user=Depends(require_admin)):
    sb = get_supabase()
    data = t.dict()
    data["estado"] = "revisar"
    res = sb.table("tendencias").insert(data).execute()
    return res.data[0]


@router.patch("/{tendencia_id}/estado")
async def cambiar_estado(
    tendencia_id: str,
    body: dict,
    user=Depends(require_admin),
):
    estado = body.get("estado")
    if estado not in ("publicada", "revisar", "rechazada"):
        raise HTTPException(status_code=400, detail="Estado inválido")

    sb = get_supabase()
    res = (
        sb.table("tendencias")
        .update({"estado": estado})
        .eq("id", tendencia_id)
        .execute()
    )
    return res.data[0]


@router.delete("/{tendencia_id}", status_code=204)
async def eliminar_tendencia(tendencia_id: str, user=Depends(require_admin)):
    sb = get_supabase()
    # Eliminar scores asociados primero
    sb.table("scores").delete().eq("tendencia_id", tendencia_id).execute()
    # Eliminar la tendencia
    res = sb.table("tendencias").delete().eq("id", tendencia_id).execute()
    if not res.data:
        raise HTTPException(status_code=404, detail="Tendencia no encontrada")


@router.post("/upload-csv")
async def upload_csv(file: UploadFile = File(...), user=Depends(require_admin)):
    if not file.filename.endswith(".csv"):
        raise HTTPException(status_code=400, detail="El archivo debe ser CSV")

    content = await file.read()
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))

    required = {"nombre", "fuente", "categoria", "descripcion"}
    rows = []

    for i, row in enumerate(reader, 1):
        if not required.issubset(row.keys()):
            raise HTTPException(
                status_code=400,
                detail=f"Columnas requeridas: {required}. Fila {i} incompleta.",
            )
        rows.append({
            "nombre": row["nombre"].strip(),
            "fuente": row["fuente"].strip(),
            "categoria": row["categoria"].strip(),
            "descripcion": row["descripcion"].strip(),
            "fecha": row.get("fecha", "").strip() or None,
            "estado": "publicada",
        })

    if not rows:
        raise HTTPException(status_code=400, detail="El CSV está vacío")

    sb = get_supabase()
    res = sb.table("tendencias").insert(rows).execute()

    return {
        "mensaje": f"{len(res.data)} tendencias cargadas correctamente",
        "tendencias": res.data,
    }


@router.post("/upload-xlsx")
async def upload_xlsx(file: UploadFile = File(...), user=Depends(require_admin)):
    if not file.filename.endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="El archivo debe ser .xlsx")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    if "Tendencias" not in wb.sheetnames or "Scoring Manual" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="El archivo debe tener hojas 'Tendencias' y 'Scoring Manual'")

    sb = get_supabase()

    # ── Leer tendencias ───────────────────────────────────────────────────────
    ws_tend = wb["Tendencias"]
    tendencias_rows = []
    for row in ws_tend.iter_rows(min_row=2, values_only=True):
        num, nombre, fuente, fecha, descripcion, categoria = (list(row) + [None]*6)[:6]
        if not isinstance(num, int) or not nombre:
            continue  # saltar filas vacías o de notas al pie
        fecha_str = fecha.strftime("%Y-%m-%d") if hasattr(fecha, "strftime") else None
        tendencias_rows.append({
            "nombre": str(nombre).strip(),
            "fuente": str(fuente).strip() if fuente else "",
            "categoria": str(categoria).strip() if categoria else "",
            "descripcion": str(descripcion).strip() if descripcion else "",
            "fecha": fecha_str,
            "estado": "publicada",
        })

    if not tendencias_rows:
        raise HTTPException(status_code=400, detail="No se encontraron tendencias en la hoja 'Tendencias'")

    # Buscar tendencias ya existentes por nombre
    nombres = [r["nombre"] for r in tendencias_rows]
    existing_res = sb.table("tendencias").select("id,nombre").in_("nombre", nombres).execute()
    existing_map = {t["nombre"]: t["id"] for t in existing_res.data}

    # Solo insertar las que no existen aún
    nuevas = [r for r in tendencias_rows if r["nombre"] not in existing_map]
    if nuevas:
        ins_res = sb.table("tendencias").insert(nuevas).execute()
        for t in ins_res.data:
            existing_map[t["nombre"]] = t["id"]

    tendencias_guardadas = existing_map

    # ── Obtener IDs de marcas por nombre ──────────────────────────────────────
    marcas_res = sb.table("marcas").select("id,nombre").execute()
    marca_ids = {m["nombre"]: m["id"] for m in marcas_res.data}

    MARCAS_XLSX = [
        {"nombre": "Norkys",      "cols": (2, 3, 4, 5)},
        {"nombre": "Papa John's", "cols": (6, 7, 8, 9)},
        {"nombre": "Starbucks",   "cols": (10, 11, 12, 13)},
    ]

    # ── Leer scores ───────────────────────────────────────────────────────────
    ws_score = wb["Scoring Manual"]
    scores_rows = []
    for row in ws_score.iter_rows(min_row=2, values_only=True):
        vals = list(row)
        num = vals[0]
        nombre = vals[1]
        if not isinstance(num, int) or not nombre:
            continue

        nombre_str = str(nombre).strip()
        tendencia_id = tendencias_guardadas.get(nombre_str)
        if not tendencia_id:
            continue

        for m in MARCAS_XLSX:
            c0, c1, c2, c3 = m["cols"]
            marca_id = marca_ids.get(m["nombre"])
            if not marca_id:
                continue
            pub  = vals[c0] or 0
            terr = vals[c1] or 0
            tono = vals[c2] or 0
            final = vals[c3] or round((pub*40 + terr*35 + tono*25) / 10, 1)
            activar = float(final) >= 65
            razon = (
                f"Score manual: Público={pub}/10, Territorio={terr}/10, Tono={tono}/10. "
                f"{'Encaje cultural suficiente para activar.' if activar else 'Encaje insuficiente para activar.'}"
            )
            scores_rows.append({
                "tendencia_id": tendencia_id,
                "marca_id": marca_id,
                "score_publico": int(pub),
                "score_territorio": int(terr),
                "score_tono": int(tono),
                "score_final": float(final),
                "activar": activar,
                "razon": razon,
                "brief": None,
            })

    if scores_rows:
        sb.table("scores").upsert(
            scores_rows, on_conflict="tendencia_id,marca_id"
        ).execute()

    return {
        "mensaje": f"{len(tendencias_guardadas)} tendencias y {len(scores_rows)} scores cargados correctamente",
        "tendencias": len(tendencias_guardadas),
        "scores": len(scores_rows),
    }
